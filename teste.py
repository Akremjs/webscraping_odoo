import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill

link = input("Entrez l'URL SVP: ")

# Envoyer une requête HTTP pour obtenir le contenu de la page en HTML
page = requests.get(link)

if page.status_code == 200:
    src = page.content
    soup = BeautifulSoup(src, "html.parser")

    # Initialiser les variables
    moduleName_text = "module name not found"
    categori_text = "categorie not found"
    odooVersion_text = "odoo version not found"
    technicalModelName_text = "technical model not found"
    price_text = "price not found"

    # Collecter les informations
    try:
        moduleName = soup.find_all(class_='mt0 mb0')[0]
        moduleName_text = moduleName.find('b').string.strip()
        print(moduleName_text)
    except (AttributeError, IndexError):
        print("Module name not found")

    try:
        categori = soup.find_all(class_='breadcrumb-item')[1]
        categori_text = categori.find('a').string.strip()
        print(categori_text)
    except (AttributeError, IndexError):
        print("Categorie not found")

    try:
        odooVersion = soup.find_all(class_='breadcrumb-item active d-flex')[0]
        odooVersion_text = odooVersion.find('span').string.strip()
        print(odooVersion_text)
    except (AttributeError, IndexError):
        print("Odoo version not found")

    try:
        technicalModelName = soup.find_all(class_='loempia_app_table table table-sm small mt16')[1]
        technicalModelName_text = technicalModelName.find('code').string.strip()
        print(technicalModelName_text)
    except (AttributeError, IndexError):
        print("Technical model not found")

    try:
        price = soup.find_all(class_='mt0 mb0')[1]
        price_text = "$ " + price.find('span').string.strip()
        print(price_text)
    except (AttributeError, IndexError):
        print("Price not found")

    # Créer un fichier Excel
    fileName = 'odoo_module_info.xlsx'

    # Créer un nouveau fichier ou charger l'existant
    try:
        wb = openpyxl.load_workbook(fileName)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'odoo_testefinal'

    # Ajouter un en-tête si c'est une nouvelle feuille
    if ws.max_row == 1 and ws.max_column == 1:
        header = ["Module Name", "Categorie", "Odoo Version", "Technical Model", "Price"]
        ws.append(header)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Ajouter les nouvelles données
    ws.append([moduleName_text, categori_text, odooVersion_text, technicalModelName_text, price_text])

    # Sauvegarder le fichier
    wb.save(fileName)
    print(f"Les données ont été enregistrées dans '{fileName}'.")

else:
    print(f"Échec de la récupération des données. Code de statut : {page.status_code}")
